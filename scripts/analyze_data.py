"""Разведка данных кейса (Блок 0): структура папок, форматы, образцы содержимого.

Запуск на ПК (Windows/Linux), Python 3.10+:
    python analyze_data.py "C:\\path\\to\\данные_кейса"

Результат — файл knot_data_report.md рядом со скриптом (и краткий вывод в консоль).
Зависимости не обязательны: чистый stdlib. Если установлены pymupdf/openpyxl —
дополнительно проверит текстовый слой PDF и листы XLSX.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

MAX_SAMPLES_PER_EXT = 4      # сколько файлов каждого типа смотреть в каждой папке
PREVIEW_CHARS = 500          # сколько символов текста показывать
CSV_ROW_CAP = 200_000        # предел строк при подсчёте


def human(n: float) -> str:
    for u in ("Б", "КБ", "МБ", "ГБ"):
        if n < 1024:
            return f"{n:.1f} {u}"
        n /= 1024
    return f"{n:.1f} ТБ"


def read_text(p: Path, limit: int = 200_000) -> str:
    data = p.read_bytes()[:limit]
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def sample_csv(p: Path) -> list[str]:
    out = []
    text = read_text(p, 2_000_000)
    try:
        dialect = csv.Sniffer().sniff(text[:4000], delimiters=",;\t|")
        delim = dialect.delimiter
    except csv.Error:
        delim = ";" if text.count(";") > text.count(",") else ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    if not rows:
        return ["  (пустой файл)"]
    out.append(f"  разделитель: `{delim!r}` · колонок: {len(rows[0])} · строк (в первых 2 МБ): {min(len(rows), CSV_ROW_CAP)}")
    out.append(f"  заголовки: {rows[0]}")
    for r in rows[1:3]:
        out.append(f"  пример: {r}")
    return out


def sample_json(p: Path) -> list[str]:
    try:
        data = json.loads(read_text(p, 2_000_000))
    except Exception as e:
        return [f"  (не парсится: {e})"]
    if isinstance(data, dict):
        return [f"  dict, ключи: {list(data.keys())[:15]}"]
    if isinstance(data, list):
        head = data[0] if data else None
        keys = list(head.keys())[:15] if isinstance(head, dict) else type(head).__name__
        return [f"  list из {len(data)} элементов, первый: {keys}"]
    return [f"  {type(data).__name__}"]


def sample_docx(p: Path) -> list[str]:
    try:
        with zipfile.ZipFile(p) as z:
            xml = z.read("word/document.xml").decode("utf-8", errors="replace")
        text = re.sub(r"<[^>]+>", " ", xml)
        text = re.sub(r"\s+", " ", text).strip()
        return [f"  текст ({len(text)} симв.): {text[:PREVIEW_CHARS]}"]
    except Exception as e:
        return [f"  (ошибка чтения docx: {e})"]


def sample_xlsx(p: Path) -> list[str]:
    try:
        import openpyxl  # noqa

        wb = openpyxl.load_workbook(p, read_only=True)
        out = [f"  листы: {wb.sheetnames}"]
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(max_row=3, values_only=True)):
            out.append(f"  строка {i + 1}: {row}")
        wb.close()
        return out
    except ImportError:
        try:
            with zipfile.ZipFile(p) as z:
                xml = z.read("xl/workbook.xml").decode("utf-8", errors="replace")
            names = re.findall(r'name="([^"]+)"', xml)
            return [f"  листы: {names} (поставь openpyxl для содержимого)"]
        except Exception as e:
            return [f"  (ошибка: {e})"]
    except Exception as e:
        return [f"  (ошибка чтения xlsx: {e})"]


def sample_pdf(p: Path) -> list[str]:
    try:
        import fitz  # pymupdf
    except ImportError:
        return ["  (поставь pymupdf для проверки текстового слоя: pip install pymupdf)"]
    try:
        doc = fitz.open(p)
        chars = [len(page.get_text()) for page in doc[: min(3, len(doc))]]
        avg = sum(chars) / max(len(chars), 1)
        verdict = "ТЕКСТОВЫЙ" if avg >= 100 else "похоже СКАН (нужен OCR!)"
        out = [f"  страниц: {len(doc)} · симв./стр (первые 3): {chars} → {verdict}"]
        if avg >= 100:
            text = re.sub(r"\s+", " ", doc[0].get_text()).strip()
            out.append(f"  начало: {text[:PREVIEW_CHARS]}")
        doc.close()
        return out
    except Exception as e:
        return [f"  (ошибка чтения pdf: {e})"]


def sample_text(p: Path) -> list[str]:
    text = re.sub(r"\s+", " ", read_text(p, 50_000)).strip()
    return [f"  начало: {text[:PREVIEW_CHARS]}"]


SAMPLERS = {
    ".csv": sample_csv, ".json": sample_json, ".docx": sample_docx,
    ".xlsx": sample_xlsx, ".xls": sample_xlsx, ".pdf": sample_pdf,
    ".md": sample_text, ".txt": sample_text,
}


def main():
    ap = argparse.ArgumentParser(description="Разведка данных кейса")
    ap.add_argument("root", help="папка с данными")
    ap.add_argument("-o", "--out", default="knot_data_report.md")
    args = ap.parse_args()
    root = Path(args.root)
    if not root.is_dir():
        sys.exit(f"Не папка: {root}")

    lines = [f"# Разведка данных: {root}", ""]

    # --- дерево верхнего уровня + статистика по папкам ---
    all_files = [p for p in root.rglob("*") if p.is_file()]
    total = sum(p.stat().st_size for p in all_files)
    lines += [f"Всего файлов: **{len(all_files)}** · объём: **{human(total)}**", "", "## Папки верхнего уровня", ""]
    top_dirs = sorted([d for d in root.iterdir() if d.is_dir()])
    per_dir_files: dict[Path, list[Path]] = defaultdict(list)
    for p in all_files:
        rel = p.relative_to(root)
        top = root / rel.parts[0] if len(rel.parts) > 1 else root
        per_dir_files[top].append(p)
    for d in top_dirs + ([root] if per_dir_files.get(root) else []):
        files = per_dir_files.get(d, [])
        size = sum(p.stat().st_size for p in files)
        exts = Counter(p.suffix.lower() or "(без расш.)" for p in files)
        name = d.name if d != root else "(файлы в корне)"
        lines.append(f"- **{name}** — {len(files)} файлов, {human(size)} · " +
                     ", ".join(f"{e}: {n}" for e, n in exts.most_common(8)))
    lines += ["", "## Расширения по всему корпусу", ""]
    for e, n in Counter(p.suffix.lower() or "(без расш.)" for p in all_files).most_common(20):
        lines.append(f"- {e}: {n}")

    # --- образцы содержимого по каждой папке ---
    for d in top_dirs + ([root] if per_dir_files.get(root) else []):
        files = per_dir_files.get(d, [])
        if not files:
            continue
        name = d.name if d != root else "(корень)"
        lines += ["", f"## Образцы: {name}", ""]
        by_ext: dict[str, list[Path]] = defaultdict(list)
        for p in sorted(files):
            by_ext[p.suffix.lower()].append(p)
        for ext, plist in sorted(by_ext.items()):
            sampler = SAMPLERS.get(ext)
            for p in plist[:MAX_SAMPLES_PER_EXT]:
                lines.append(f"### {p.relative_to(root)} ({human(p.stat().st_size)})")
                lines += (sampler(p) if sampler else ["  (тип не сэмплируем)"])
                lines.append("")

    report = "\n".join(lines) + "\n"
    out_path = Path(args.out)
    out_path.write_text(report, encoding="utf-8")
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    print(report[:6000])
    print(f"\n=== Полный отчёт: {out_path.resolve()} — пришли этот файл в чат ===")


if __name__ == "__main__":
    main()
